from __future__ import annotations

import csv
import os
import re
import subprocess
import tempfile
import time
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any

from google.auth.transport.requests import Request
from google.oauth2.credentials import Credentials
from googleapiclient.discovery import build
from openpyxl import load_workbook


SCOPES = [
    "https://www.googleapis.com/auth/gmail.modify",
    "https://www.googleapis.com/auth/gmail.send",
    "https://www.googleapis.com/auth/drive",
    "https://www.googleapis.com/auth/drive.readonly",
    "https://www.googleapis.com/auth/spreadsheets",
]

DEFAULT_CREDENTIALS_PATH = "/home/brubot77/email-handler-emily/credentials.json"
DEFAULT_TOKEN_PATH = "/home/brubot77/email-handler-emily/token.json"

NATIVE_SHEET_NAME = "BLU Active Deals - Google Sheet"
ACTIVE_DEALS_TAB_NAMES = {"active deals", "blu active deals"}

SHANNON_DIR = Path("/home/brubot77/.openclaw/workspace/shannon")
SHANNON_INPUT_DIR = SHANNON_DIR / "Input"
SHANNON_OUTPUT_DIR = SHANNON_DIR / "Output"


@dataclass
class ActiveDealsResult:
    action: str
    file_name: str
    row_number: int
    address: str
    match_key: str
    updated_fields: list[str]


def _norm(value: Any) -> str:
    return re.sub(r"[^a-z0-9]+", "", str(value or "").strip().lower())


def _a1_col(col_idx: int) -> str:
    letters = ""

    while col_idx:
        col_idx, remainder = divmod(col_idx - 1, 26)
        letters = chr(65 + remainder) + letters

    return letters


def _quote_sheet_name(sheet_name: str) -> str:
    return "'" + sheet_name.replace("'", "''") + "'"


def _a1(sheet_name: str, row_idx: int, col_idx: int) -> str:
    return f"{_quote_sheet_name(sheet_name)}!{_a1_col(col_idx)}{row_idx}"


def normalize_address_match_key(address: str) -> str:
    text = str(address or "").strip().lower()

    if "," in text:
        text = text.split(",", 1)[0]

    text = re.sub(r"\bwichita\b|\bks\b|\bkansas\b", " ", text, flags=re.IGNORECASE)
    text = re.sub(r"\b\d{5}(?:-\d{4})?\b", " ", text)
    text = re.sub(r"[^a-z0-9\s]", " ", text)

    directions = {
        "n", "s", "e", "w",
        "ne", "nw", "se", "sw",
        "north", "south", "east", "west",
        "northeast", "northwest", "southeast", "southwest",
    }

    suffixes = {
        "street", "st",
        "avenue", "ave",
        "road", "rd",
        "drive", "dr",
        "lane", "ln",
        "court", "ct",
        "place", "pl",
        "boulevard", "blvd",
        "terrace", "ter",
        "parkway", "pkwy",
        "circle", "cir",
        "trail", "trl",
        "way",
        "loop",
        "highway", "hwy",
        "route",
    }

    unit_words = {
        "apt", "apartment", "unit", "suite", "ste",
    }

    words = []
    skip_next = False

    for word in text.split():
        if skip_next:
            skip_next = False
            continue

        if word in unit_words:
            skip_next = True
            continue

        if word in directions:
            continue

        if word in suffixes:
            continue

        words.append(word)

    if not words:
        return ""

    number = next((w for w in words if w.isdigit()), "")

    if not number:
        return " ".join(words[:2]).strip()

    number_idx = words.index(number)

    street_name = ""
    for word in words[number_idx + 1:]:
        if not word.isdigit():
            street_name = word
            break

    if street_name:
        return f"{number} {street_name}".strip()

    return number


def _connect_services():
    credentials_path = Path(os.getenv("ACTIVE_DEALS_CREDENTIALS_PATH", DEFAULT_CREDENTIALS_PATH))
    token_path = Path(os.getenv("ACTIVE_DEALS_TOKEN_PATH", DEFAULT_TOKEN_PATH))

    if not credentials_path.exists():
        raise FileNotFoundError(f"Google credentials file not found: {credentials_path}")

    if not token_path.exists():
        raise FileNotFoundError(f"Google token file not found: {token_path}")

    creds = Credentials.from_authorized_user_file(
        str(token_path),
        scopes=SCOPES,
    )

    if not creds.valid:
        if creds.expired and creds.refresh_token:
            creds.refresh(Request())
            token_path.write_text(creds.to_json(), encoding="utf-8")
        else:
            raise RuntimeError("Google token is invalid. Re-run OAuth with Sheets scope.")

    drive = build("drive", "v3", credentials=creds)
    sheets = build("sheets", "v4", credentials=creds)

    return drive, sheets


def find_native_active_deals_sheet() -> tuple[str, str]:
    drive, _sheets = _connect_services()

    response = drive.files().list(
        q=(
            f"name = '{NATIVE_SHEET_NAME}' "
            "and mimeType = 'application/vnd.google-apps.spreadsheet' "
            "and trashed = false"
        ),
        spaces="drive",
        fields="files(id,name,mimeType,modifiedTime)",
        includeItemsFromAllDrives=True,
        supportsAllDrives=True,
        pageSize=10,
    ).execute()

    files = response.get("files", [])

    if not files:
        raise FileNotFoundError(f"No native Google Sheet found named {NATIVE_SHEET_NAME}")

    return files[0]["id"], files[0]["name"]


def _get_spreadsheet_meta(sheets, spreadsheet_id: str) -> dict:
    return sheets.spreadsheets().get(
        spreadsheetId=spreadsheet_id,
        fields="sheets(properties(sheetId,title,index,gridProperties),protectedRanges(protectedRangeId,description,range))",
    ).execute()


def _get_active_sheet_props(meta: dict) -> dict:
    for sheet in meta.get("sheets", []):
        props = sheet["properties"]

        if props["title"].strip().lower() in ACTIVE_DEALS_TAB_NAMES:
            return props

    raise ValueError("Could not find Active Deals tab in native Google Sheet")


def _get_sheet_props_by_title(meta: dict, title: str) -> dict | None:
    for sheet in meta.get("sheets", []):
        props = sheet["properties"]

        if props["title"] == title:
            return props

    return None


def _read_values(sheets, spreadsheet_id: str, range_name: str) -> list[list[Any]]:
    return sheets.spreadsheets().values().get(
        spreadsheetId=spreadsheet_id,
        range=range_name,
    ).execute().get("values", [])


def _find_header_row_and_map(values: list[list[Any]]) -> tuple[int, dict[str, int]]:
    for idx, row in enumerate(values, start=1):
        normalized = [_norm(value) for value in row]

        if "address" in normalized:
            headers = {}

            for col_idx, raw_header in enumerate(row, start=1):
                key = _norm(raw_header)

                if key:
                    headers[key] = col_idx

            return idx, headers

    raise ValueError("Could not find header row with Address column")


def _find_col(headers: dict[str, int], candidates: list[str]) -> int | None:
    normalized = [_norm(c) for c in candidates]

    for candidate in normalized:
        if candidate in headers:
            return headers[candidate]

    for key, col_idx in headers.items():
        if any(candidate in key for candidate in normalized):
            return col_idx

    return None


def _prepare_active_deals_text(body_text: str) -> str:
    text = (body_text or "").replace("\r\n", "\n").replace("\r", "\n").strip()

    known_labels = [
        "address",
        "property address",
        "street address",
        "property",
        "status",
        "status update",
        "offer status",
        "rent",
        "monthly rent",
        "zest rent",
        "market rent",
        "blu rent est",
        "blue rent est",
        "price",
        "asking price",
        "seller price",
        "latest offer",
        "purchase price",
        "offer price",
        "appraisal est",
        "appraisal estimate",
        "rehab est",
        "rehab",
        "beds",
        "bedrooms",
        "baths",
        "bathrooms",
        "source",
        "notes",
        "note",
        "comments",
        "comment",
    ]

    label_pattern = "|".join(re.escape(label) for label in known_labels)

    text = re.sub(
        rf"(?i)(?<!^)(?<!\n)({label_pattern})\s*:",
        r"\n\1:",
        text,
    )

    return text


def _extract_labeled_fields(body_text: str) -> dict[str, str]:
    fields: dict[str, str] = {}
    text = _prepare_active_deals_text(body_text)

    pattern = re.compile(r"(?im)^\s*([A-Za-z][A-Za-z0-9 /&()._-]{1,50})\s*:\s*(.*)$")
    matches = list(pattern.finditer(text))

    for i, match in enumerate(matches):
        label = match.group(1).strip()
        start_value = match.group(2).strip()

        next_start = matches[i + 1].start() if i + 1 < len(matches) else len(text)
        value_tail_start = match.end()
        tail = text[value_tail_start:next_start].strip()

        if tail:
            value = f"{start_value}\n{tail}".strip()
        else:
            value = start_value

        value = re.sub(r"\n+", " ", value)
        value = re.sub(r"\s+", " ", value).strip()

        if label:
            fields[_norm(label)] = value

    return fields


def _looks_like_address_line(line: str) -> bool:
    line = str(line or "").strip()

    if not line:
        return False

    if ":" in line:
        return False

    return bool(
        re.search(
            r"^\s*\d{2,6}\s+(?:[NSEW]\.?\s+|North\s+|South\s+|East\s+|West\s+)?[A-Za-z0-9.'-]+(?:\s+[A-Za-z0-9.'-]+){0,4}(?:,\s*[A-Za-z .'-]+)?",
            line,
            re.IGNORECASE,
        )
    )


def _extract_address_from_body(body_text: str, fields: dict[str, str]) -> str:
    for key in ["address", "propertyaddress", "streetaddress", "property"]:
        value = fields.get(key)

        if value:
            return value

    text = _prepare_active_deals_text(body_text)

    for line in text.splitlines():
        line = line.strip()

        if _looks_like_address_line(line):
            return line

    return ""


def parse_active_deal_blocks(body_text: str) -> list[str]:
    text = _prepare_active_deals_text(body_text)

    if not text:
        return []

    lines = text.splitlines()

    blocks: list[list[str]] = []
    current: list[str] = []

    for raw_line in lines:
        line = raw_line.strip()

        if not line:
            if current:
                current.append("")
            continue

        is_address = _looks_like_address_line(line) or line.lower().startswith("address:")

        if is_address and current:
            cleaned_current = "\n".join(current).strip()

            if cleaned_current:
                blocks.append(current)

            current = [line]
        else:
            current.append(line)

    if current:
        cleaned_current = "\n".join(current).strip()

        if cleaned_current:
            blocks.append(current)

    return ["\n".join(block).strip() for block in blocks if "\n".join(block).strip()]


def _field_aliases_for_header(header_key: str) -> list[str]:
    aliases = {
        "address": ["address", "property", "propertyaddress", "streetaddress"],
        "city": ["city"],
        "st": ["state", "st"],
        "state": ["state", "st"],
        "sellerprice": ["sellerprice", "price", "askingprice"],
        "appraisalest": ["appraisalest", "appraisalestimate", "appraisal"],
        "latestoffer": ["latestoffer", "offer", "offerprice", "purchaseprice"],
        "rehabest": ["rehabest", "rehab", "rehabbudget", "repairs"],
        "offerstatus": ["offerstatus", "status", "statusupdate"],
        "propertystatus": ["status", "offerstatus"],
        "propertynotes": ["propertynotes", "notes", "note", "comments", "comment"],
        "annualtax": ["annualtax", "tax", "taxes"],
        "insuranceest": ["insuranceest", "insurance"],
        "rent": ["rent", "monthlyrent", "zestrent", "marketrent", "blurentest", "bluerentest"],
    }

    return aliases.get(header_key, [header_key])


def _split_city_state_zip(address: str) -> tuple[str, str, str]:
    parts = [p.strip() for p in str(address or "").split(",")]

    city = ""
    state = ""
    zip_code = ""

    if len(parts) >= 2:
        city = parts[1].strip()

    if len(parts) >= 3:
        state_zip = parts[2].strip()
        tokens = state_zip.split()

        if tokens:
            state = tokens[0].strip()

        if len(tokens) > 1:
            zip_code = tokens[1].strip()

    return city, state, zip_code


def _existing_match_row(values: list[list[Any]], header_row: int, address_col: int, incoming_key: str) -> int | None:
    for row_idx in range(header_row + 1, len(values) + 1):
        row = values[row_idx - 1]
        existing_address = row[address_col - 1] if len(row) >= address_col else ""
        existing_key = normalize_address_match_key(existing_address)

        if existing_key and existing_key == incoming_key:
            return row_idx

    return None


def _first_empty_address_row(values: list[list[Any]], header_row: int, address_col: int) -> int:
    for row_idx in range(header_row + 1, len(values) + 1):
        row = values[row_idx - 1]
        value = row[address_col - 1] if len(row) >= address_col else ""

        if str(value or "").strip() == "":
            return row_idx

    return len(values) + 1


def _safe_sheet_title(value: str) -> str:
    title = str(value or "").strip()

    if "," in title:
        title = title.split(",", 1)[0].strip()

    title = re.sub(r"[:\\/?*\[\]]", " ", title)
    title = re.sub(r"\s+", " ", title).strip()

    if not title:
        title = "Property"

    return title[:31]


def _unique_sheet_title(meta: dict, desired_title: str) -> str:
    desired_title = _safe_sheet_title(desired_title)

    existing = {s["properties"]["title"] for s in meta.get("sheets", [])}

    if desired_title not in existing:
        return desired_title

    base = desired_title[:27].rstrip()
    counter = 2

    while True:
        candidate = f"{base} {counter}"[:31]

        if candidate not in existing:
            return candidate

        counter += 1


def _sheet_exists_for_address(meta: dict, address: str) -> bool:
    incoming_key = normalize_address_match_key(address)

    if not incoming_key:
        return False

    for sheet in meta.get("sheets", []):
        title = sheet["properties"]["title"]
        sheet_key = normalize_address_match_key(title)

        if sheet_key and sheet_key == incoming_key:
            return True

    return False


def _split_address_city_state(address: str) -> tuple[str, str, str]:
    parts = [p.strip() for p in str(address or "").split(",")]

    street = parts[0] if parts else str(address or "").strip()
    city = parts[1] if len(parts) >= 2 else "Wichita"
    state = "KS"

    if len(parts) >= 3:
        state_part = parts[2].strip().split()
        if state_part:
            state = state_part[0]

    return street, city, state


def _snapshot_shannon_outputs() -> dict[str, float]:
    return {
        str(path): path.stat().st_mtime
        for path in SHANNON_OUTPUT_DIR.glob("*.xlsx")
    }


def _newest_shannon_output_after(before_snapshot: dict[str, float]) -> Path:
    candidates = sorted(
        SHANNON_OUTPUT_DIR.glob("*.xlsx"),
        key=lambda p: p.stat().st_mtime,
        reverse=True,
    )

    for path in candidates:
        old_mtime = before_snapshot.get(str(path))

        if old_mtime is None:
            return path

        if path.stat().st_mtime > old_mtime:
            return path

    raise FileNotFoundError("Shannon ran but no new output workbook was detected")


def _run_shannon_for_property(
    address: str,
    city: str = "",
    state: str = "",
) -> Path:
    street, parsed_city, parsed_state = _split_address_city_state(address)
    city = str(city or "").strip() or parsed_city
    state = str(state or "").strip() or parsed_state or "KS"

    print(
        f"Active Deals -> Shannon input: address={street!r}, "
        f"city={city!r}, state={state!r}"
    )

    SHANNON_INPUT_DIR.mkdir(parents=True, exist_ok=True)

    stamp = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
    safe_name = re.sub(r"[^A-Za-z0-9]+", "_", street).strip("_") or "property"
    csv_path = SHANNON_INPUT_DIR / f"active_deals_{safe_name}_{stamp}.csv"

    with csv_path.open("w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(
            f,
            fieldnames=["Address", "City", "State"],
        )
        writer.writeheader()
        writer.writerow(
            {
                "Address": street,
                "City": city,
                "State": state,
            }
        )

    before_snapshot = _snapshot_shannon_outputs()

    cmd = (
        f"cd {SHANNON_DIR} "
        f"&& source .venv/bin/activate "
        f"&& python3 -m shannon.cli"
    )

    result = subprocess.run(
        ["bash", "-lc", cmd],
        capture_output=True,
        text=True,
        timeout=240,
    )

    if result.stdout:
        print(result.stdout)

    if result.stderr:
        print(result.stderr)

    if result.returncode != 0:
        raise RuntimeError(f"Shannon failed for {address}")

    time.sleep(1)

    return _newest_shannon_output_after(before_snapshot)


def _shannon_workbook_to_values(shannon_output_path: Path) -> list[list[Any]]:
    wb = load_workbook(shannon_output_path, data_only=False)
    ws = wb[wb.sheetnames[0]]

    values: list[list[Any]] = []

    for row in range(1, ws.max_row + 1):
        row_values = []

        for col in range(1, ws.max_column + 1):
            value = ws.cell(row=row, column=col).value

            if value is None:
                value = ""

            row_values.append(value)

        values.append(row_values)

    return values


def _add_sheet(sheets, spreadsheet_id: str, title: str) -> int:
    response = sheets.spreadsheets().batchUpdate(
        spreadsheetId=spreadsheet_id,
        body={
            "requests": [
                {
                    "addSheet": {
                        "properties": {
                            "title": title,
                            "gridProperties": {
                                "rowCount": 100,
                                "columnCount": 70,
                            },
                        }
                    }
                }
            ]
        },
    ).execute()

    return response["replies"][0]["addSheet"]["properties"]["sheetId"]


def _write_values(sheets, spreadsheet_id: str, range_name: str, values: list[list[Any]]) -> None:
    sheets.spreadsheets().values().update(
        spreadsheetId=spreadsheet_id,
        range=range_name,
        valueInputOption="USER_ENTERED",
        body={"values": values},
    ).execute()

def _execute_with_retry(request, description: str = "Google API request"):
    """
    Handle temporary Google API rate limits.

    Main fix is batching, but this gives us a safety net.
    """
    for attempt in range(6):
        try:
            return request.execute()
        except HttpError as exc:
            status = getattr(exc.resp, "status", None)

            if status in {429, 500, 502, 503, 504} and attempt < 5:
                sleep_seconds = min(60, 5 * (attempt + 1))
                print(f"{description} rate-limited; sleeping {sleep_seconds}s then retrying")
                time.sleep(sleep_seconds)
                continue

            raise


def _write_values_batch(
    sheets,
    spreadsheet_id: str,
    updates: list[tuple[str, Any]],
) -> None:
    """
    Batch multiple cell formula/value updates into one Sheets API write request.
    """

    if not updates:
        return

    data = [
        {
            "range": range_name,
            "values": [[value]],
        }
        for range_name, value in updates
    ]

    request = sheets.spreadsheets().values().batchUpdate(
        spreadsheetId=spreadsheet_id,
        body={
            "valueInputOption": "USER_ENTERED",
            "data": data,
        },
    )

    _execute_with_retry(request, "Sheets values batchUpdate")


def _batch_sheet_requests(
    sheets,
    spreadsheet_id: str,
    requests: list[dict],
    description: str = "Sheets batchUpdate",
) -> None:
    """
    Batch format/hide/protection requests into one Sheets API write request.
    """

    if not requests:
        return

    request = sheets.spreadsheets().batchUpdate(
        spreadsheetId=spreadsheet_id,
        body={"requests": requests},
    )

    _execute_with_retry(request, description)


def _format_column_request(
    sheet_id: int,
    col_1based: int,
    number_format_type: str,
    pattern: str,
) -> dict:
    return {
        "repeatCell": {
            "range": {
                "sheetId": sheet_id,
                "startRowIndex": 0,
                "endRowIndex": 1000,
                "startColumnIndex": col_1based - 1,
                "endColumnIndex": col_1based,
            },
            "cell": {
                "userEnteredFormat": {
                    "numberFormat": {
                        "type": number_format_type,
                        "pattern": pattern,
                    }
                }
            },
            "fields": "userEnteredFormat.numberFormat",
        }
    }


def _format_cell_request(
    sheet_id: int,
    row_1based: int,
    col_1based: int,
    number_format_type: str,
    pattern: str,
) -> dict:
    return {
        "repeatCell": {
            "range": {
                "sheetId": sheet_id,
                "startRowIndex": row_1based - 1,
                "endRowIndex": row_1based,
                "startColumnIndex": col_1based - 1,
                "endColumnIndex": col_1based,
            },
            "cell": {
                "userEnteredFormat": {
                    "numberFormat": {
                        "type": number_format_type,
                        "pattern": pattern,
                    }
                }
            },
            "fields": "userEnteredFormat.numberFormat",
        }
    }


def _hide_column_request(sheet_id: int, col_1based: int) -> dict | None:
    """
    Never hide A or B.
    """

    if col_1based <= 2:
        return None

    return {
        "updateDimensionProperties": {
            "range": {
                "sheetId": sheet_id,
                "dimension": "COLUMNS",
                "startIndex": col_1based - 1,
                "endIndex": col_1based,
            },
            "properties": {
                "hiddenByUser": True,
            },
            "fields": "hiddenByUser",
        }
    }


def _protect_range_request(
    sheet_id: int,
    start_row_1based: int,
    end_row_1based_inclusive: int,
    start_col_1based: int,
    end_col_1based_inclusive: int,
    description: str,
) -> dict:
    return {
        "addProtectedRange": {
            "protectedRange": {
                "range": {
                    "sheetId": sheet_id,
                    "startRowIndex": start_row_1based - 1,
                    "endRowIndex": end_row_1based_inclusive,
                    "startColumnIndex": start_col_1based - 1,
                    "endColumnIndex": end_col_1based_inclusive,
                },
                "description": description,
                "warningOnly": True,
            }
        }
    }

def _delete_existing_protections_by_description(
    sheets,
    spreadsheet_id: str,
    sheet_id: int,
    description: str,
) -> None:
    meta = sheets.spreadsheets().get(
        spreadsheetId=spreadsheet_id,
        fields="sheets(properties(sheetId,title),protectedRanges(protectedRangeId,description,range))",
    ).execute()

    delete_requests = []

    for sheet in meta.get("sheets", []):
        props = sheet.get("properties", {})

        if props.get("sheetId") != sheet_id:
            continue

        for protected_range in sheet.get("protectedRanges", []) or []:
            if protected_range.get("description") == description:
                delete_requests.append(
                    {
                        "deleteProtectedRange": {
                            "protectedRangeId": protected_range["protectedRangeId"]
                        }
                    }
                )

    if delete_requests:
        sheets.spreadsheets().batchUpdate(
            spreadsheetId=spreadsheet_id,
            body={"requests": delete_requests},
        ).execute()


def _protect_range(
    sheets,
    spreadsheet_id: str,
    sheet_id: int,
    start_row_1based: int,
    end_row_1based_inclusive: int,
    start_col_1based: int,
    end_col_1based_inclusive: int,
    description: str,
) -> None:
    """
    Add native Google Sheets protection.

    warningOnly=True is intentional because owners can usually still edit hard
    protected ranges. Warning protection gives a clear accidental-edit guard
    in Google Sheets.
    """

    _delete_existing_protections_by_description(
        sheets=sheets,
        spreadsheet_id=spreadsheet_id,
        sheet_id=sheet_id,
        description=description,
    )

    sheets.spreadsheets().batchUpdate(
        spreadsheetId=spreadsheet_id,
        body={
            "requests": [
                {
                    "addProtectedRange": {
                        "protectedRange": {
                            "range": {
                                "sheetId": sheet_id,
                                "startRowIndex": start_row_1based - 1,
                                "endRowIndex": end_row_1based_inclusive,
                                "startColumnIndex": start_col_1based - 1,
                                "endColumnIndex": end_col_1based_inclusive,
                            },
                            "description": description,
                            "warningOnly": True,
                        }
                    }
                }
            ]
        },
    ).execute()

def _format_cell(
    sheets,
    spreadsheet_id: str,
    sheet_id: int,
    row_1based: int,
    col_1based: int,
    number_format_type: str,
    pattern: str,
) -> None:
    sheets.spreadsheets().batchUpdate(
        spreadsheetId=spreadsheet_id,
        body={
            "requests": [
                {
                    "repeatCell": {
                        "range": {
                            "sheetId": sheet_id,
                            "startRowIndex": row_1based - 1,
                            "endRowIndex": row_1based,
                            "startColumnIndex": col_1based - 1,
                            "endColumnIndex": col_1based,
                        },
                        "cell": {
                            "userEnteredFormat": {
                                "numberFormat": {
                                    "type": number_format_type,
                                    "pattern": pattern,
                                }
                            }
                        },
                        "fields": "userEnteredFormat.numberFormat",
                    }
                }
            ]
        },
    ).execute()

def _find_header_columns_in_values(values: list[list[Any]]) -> dict[str, int]:
    """
    Find Shannon header row and return normalized header -> 1-based column index.
    Shannon individual tabs typically have the main headers on row 15.
    """

    for row in values:
        normalized = [_norm(value) for value in row]

        if "address" in normalized and "zestrent" in normalized:
            headers: dict[str, int] = {}

            for col_idx, raw_header in enumerate(row, start=1):
                key = _norm(raw_header)

                if key:
                    headers[key] = col_idx

            return headers

    return {}


def _format_column(
    sheets,
    spreadsheet_id: str,
    sheet_id: int,
    col_1based: int,
    number_format_type: str,
    pattern: str,
) -> None:
    sheets.spreadsheets().batchUpdate(
        spreadsheetId=spreadsheet_id,
        body={
            "requests": [
                {
                    "repeatCell": {
                        "range": {
                            "sheetId": sheet_id,
                            "startRowIndex": 0,
                            "endRowIndex": 1000,
                            "startColumnIndex": col_1based - 1,
                            "endColumnIndex": col_1based,
                        },
                        "cell": {
                            "userEnteredFormat": {
                                "numberFormat": {
                                    "type": number_format_type,
                                    "pattern": pattern,
                                }
                            }
                        },
                        "fields": "userEnteredFormat.numberFormat",
                    }
                }
            ]
        },
    ).execute()


def _hide_column(
    sheets,
    spreadsheet_id: str,
    sheet_id: int,
    col_1based: int,
) -> None:
    """
    Hide a column in Google Sheets.

    Never hide A or B.
    """

    if col_1based <= 2:
        return

    sheets.spreadsheets().batchUpdate(
        spreadsheetId=spreadsheet_id,
        body={
            "requests": [
                {
                    "updateDimensionProperties": {
                        "range": {
                            "sheetId": sheet_id,
                            "dimension": "COLUMNS",
                            "startIndex": col_1based - 1,
                            "endIndex": col_1based,
                        },
                        "properties": {
                            "hiddenByUser": True,
                        },
                        "fields": "hiddenByUser",
                    }
                }
            ]
        },
    ).execute()


def _format_and_hide_property_tab_columns(
    sheets,
    spreadsheet_id: str,
    property_sheet_id: int,
    shannon_values: list[list[Any]],
) -> None:
    """
    Apply user-friendly formatting and hide less-used Shannon detail columns
    on individual property tabs using one batched Sheets API request.
    """

    headers = _find_header_columns_in_values(shannon_values)

    if not headers:
        print("Could not find Shannon headers for property tab formatting")
        return

    requests: list[dict] = []

    percent_headers = [
        "Interest Rate",
        "Maintenance %",
        "Vacancy %",
        "Closing Cost %",
        "Down Payment %",
        "Cash on Cash %",
        "Cap Rate Est",
    ]

    currency_headers = [
        "Property Mgmt $/Door (Monthly)",
        "Appraisal Fee $/Door",
        "Zest Rent",
        "Price",
        "Rent Estimate (Monthly)",
        "Rent Median",
        "Rent Mean",
        "Rent Min",
        "Annual Taxes",
        "Tax Appraisal",
        "Tax Land Value",
        "Insurance Annual",
        "Rehab Cost (Override)",
        "Closing Cost $",
        "Monthly PI Payment",
        "Monthly Maintenance $",
        "Monthly Vacancy $",
        "Monthly Mgmt $",
        "Monthly Tax $",
        "Monthly Insurance $",
        "Monthly Operating Expenses $",
        "Monthly Cashflow Est",
        "NOI Annual Est",
    ]

    for header in percent_headers:
        col_idx = _find_col(headers, [header])

        if col_idx:
            requests.append(
                _format_column_request(
                    sheet_id=property_sheet_id,
                    col_1based=col_idx,
                    number_format_type="PERCENT",
                    pattern="0.00%",
                )
            )

    for header in currency_headers:
        col_idx = _find_col(headers, [header])

        if col_idx:
            requests.append(
                _format_column_request(
                    sheet_id=property_sheet_id,
                    col_1based=col_idx,
                    number_format_type="CURRENCY",
                    pattern="$#,##0.00;[Red]($#,##0.00)",
                )
            )

    headers_to_hide = [
        "Type",
        "State",
        "Zip",
        "Rent P25",
        "Rent P75",
        "Rent Max",
        "Rent StdDev",
        "Rent Error",
        "County Tax PIN",
        "Rehab Cost (Override)",
        "Maintenance %",
        "Vacancy %",
        "Property Mgmt $/Door (Monthly)",
        "Appraisal Fee $/Door",
        "Interest Rate",
        "Amortization (Years)",
        "Closing Cost %",
        "Down Payment %",
        "Closing Cost $",
        "Monthly PI Payment",
        "Monthly Maintenance $",
        "Monthly Mgmt $",
        "Monthly Tax $",
        "Monthly Insurance $",
        "Monthly Operating Expenses $",
        "Monthly Cashflow Est",
        "NOI Annual Est",
        "Cap Rate Est",
        "Rent Cache Hit",
        "Tax Cache Hit",
        "Rent Last Checked UTC",
        "Tax Last Checked UTC",
        "Tax Year",
        "Tax Source",
        "Last Reviewed UTC",
    ]

    for header in headers_to_hide:
        col_idx = _find_col(headers, [header])

        if col_idx:
            request = _hide_column_request(
                sheet_id=property_sheet_id,
                col_1based=col_idx,
            )

            if request:
                requests.append(request)

    _batch_sheet_requests(
        sheets=sheets,
        spreadsheet_id=spreadsheet_id,
        requests=requests,
        description="Format/hide property tab columns",
    )

def _link_property_tab_formulas(
    sheets,
    spreadsheet_id: str,
    active_sheet_id: int,
    property_sheet_id: int,
    active_sheet_name: str,
    property_sheet_name: str,
    headers: dict[str, int],
    active_row_idx: int,
) -> None:
    latest_offer_col = _find_col(headers, ["Latest Offer", "Offer", "Offer Price", "Purchase Price", "Price"])
    appraisal_est_col = _find_col(headers, ["Appraisal Est.", "Appraisal Est", "Appraisal Estimate", "Appraisal"])
    rehab_est_col = _find_col(headers, ["Rehab Est.", "Rehab Est", "Rehab Estimate", "Rehab", "Repairs"])
    conc_col = _find_col(headers, ["ConC", "CoC", "Cash on Cash", "Cash on Cash %"])
    annual_cashflow_col = _find_col(headers, ["Annual Cashflow", "Annual Cash Flow", "Annual CF"])
    cash_left_col = _find_col(headers, ["Cash Left in Deal", "Cash Left In Deal", "Cash Left", "Cash In Deal"])
    
    formula_updates: list[tuple[str, Any]] = []

    if latest_offer_col:
        latest_offer_ref = _a1(active_sheet_name, active_row_idx, latest_offer_col)
        formula_updates.append((
            _a1(property_sheet_name, 16, 15),  # O16
            f"={latest_offer_ref}",
        ))

    if latest_offer_col and appraisal_est_col:
        latest_offer_ref = _a1(active_sheet_name, active_row_idx, latest_offer_col)
        appraisal_ref = _a1(active_sheet_name, active_row_idx, appraisal_est_col)

        formula_updates.append((
            _a1(property_sheet_name, 5, 2),  # B5
            f'=IFERROR(1-(0.8*{appraisal_ref}/{latest_offer_ref}),0.2)',
        ))

    # Individual tab Rehab Cost Override AG16 = Active Deals Rehab Est.
    if rehab_est_col:
        rehab_est_ref = _a1(active_sheet_name, active_row_idx, rehab_est_col)

        formula_updates.append((
            _a1(property_sheet_name, 16, 33),  # AG16
            f"={rehab_est_ref}",
        ))

    if cash_left_col:
        formula_updates.append((
            _a1(active_sheet_name, active_row_idx, cash_left_col),
            f"={_a1(property_sheet_name, 16, 15)}*({_a1(property_sheet_name, 16, 41)}+{_a1(property_sheet_name, 16, 40)})",
        ))

    if annual_cashflow_col:
        formula_updates.append((
            _a1(active_sheet_name, active_row_idx, annual_cashflow_col),
            f"=12*{_a1(property_sheet_name, 16, 50)}",
        ))

    if conc_col:
        if cash_left_col:
            cash_left_ref = _a1(active_sheet_name, active_row_idx, cash_left_col)
            formula = f'=IF({cash_left_ref}<0,"CASH OUT",{_a1(property_sheet_name, 16, 16)})'
        else:
            formula = f"={_a1(property_sheet_name, 16, 16)}"

        formula_updates.append((
            _a1(active_sheet_name, active_row_idx, conc_col),
            formula,
        ))

    _write_values_batch(
        sheets=sheets,
        spreadsheet_id=spreadsheet_id,
        updates=formula_updates,
    )

    # Native Google Sheets number formats and protected ranges.
    requests: list[dict] = []

    # Property tab:
    #   B5 = Down Payment %
    #   O16 = Price
    #   P16 = Cash on Cash %
    #   AG16 = Rehab Cost Override
    requests.append(
        _format_cell_request(
            sheet_id=property_sheet_id,
            row_1based=5,
            col_1based=2,
            number_format_type="PERCENT",
            pattern="0.00%",
        )
    )

    requests.append(
        _format_cell_request(
            sheet_id=property_sheet_id,
            row_1based=16,
            col_1based=15,
            number_format_type="CURRENCY",
            pattern="$#,##0.00;[Red]($#,##0.00)",
        )
    )

    requests.append(
        _format_cell_request(
            sheet_id=property_sheet_id,
            row_1based=16,
            col_1based=16,
            number_format_type="PERCENT",
            pattern="0.00%",
        )
    )

    requests.append(
        _format_cell_request(
            sheet_id=property_sheet_id,
            row_1based=16,
            col_1based=33,
            number_format_type="CURRENCY",
            pattern="$#,##0.00;[Red]($#,##0.00)",
        )
    )

    # Active Deals:
    #   Cash Left in Deal = $
    #   ConC = %
    #   Annual Cashflow = $
    if cash_left_col:
        requests.append(
            _format_cell_request(
                sheet_id=active_sheet_id,
                row_1based=active_row_idx,
                col_1based=cash_left_col,
                number_format_type="CURRENCY",
                pattern="$#,##0.00;[Red]($#,##0.00)",
            )
        )

    if conc_col:
        requests.append(
            _format_cell_request(
                sheet_id=active_sheet_id,
                row_1based=active_row_idx,
                col_1based=conc_col,
                number_format_type="PERCENT",
                pattern="0.00%",
            )
        )

    if annual_cashflow_col:
        requests.append(
            _format_cell_request(
                sheet_id=active_sheet_id,
                row_1based=active_row_idx,
                col_1based=annual_cashflow_col,
                number_format_type="CURRENCY",
                pattern="$#,##0.00;[Red]($#,##0.00)",
            )
        )

    # Protected formula cells.
    requests.append(
        _protect_range_request(
            sheet_id=property_sheet_id,
            start_row_1based=5,
            end_row_1based_inclusive=5,
            start_col_1based=2,
            end_col_1based_inclusive=2,
            description=f"Emily protected down payment formula for {property_sheet_name}",
        )
    )

    requests.append(
        _protect_range_request(
            sheet_id=property_sheet_id,
            start_row_1based=16,
            end_row_1based_inclusive=16,
            start_col_1based=15,
            end_col_1based_inclusive=15,
            description=f"Emily protected price formula for {property_sheet_name}",
        )
    )

    if rehab_est_col:
        requests.append(
            _protect_range_request(
                sheet_id=property_sheet_id,
                start_row_1based=16,
                end_row_1based_inclusive=16,
                start_col_1based=33,
                end_col_1based_inclusive=33,
                description=f"Emily protected rehab override formula for {property_sheet_name}",
            )
        )

    if cash_left_col:
        requests.append(
            _protect_range_request(
                sheet_id=active_sheet_id,
                start_row_1based=active_row_idx,
                end_row_1based_inclusive=active_row_idx,
                start_col_1based=cash_left_col,
                end_col_1based_inclusive=cash_left_col,
                description=f"Emily protected Cash Left formula for {property_sheet_name}",
            )
        )

    if conc_col:
        requests.append(
            _protect_range_request(
                sheet_id=active_sheet_id,
                start_row_1based=active_row_idx,
                end_row_1based_inclusive=active_row_idx,
                start_col_1based=conc_col,
                end_col_1based_inclusive=conc_col,
                description=f"Emily protected ConC formula for {property_sheet_name}",
            )
        )

    if annual_cashflow_col:
        requests.append(
            _protect_range_request(
                sheet_id=active_sheet_id,
                start_row_1based=active_row_idx,
                end_row_1based_inclusive=active_row_idx,
                start_col_1based=annual_cashflow_col,
                end_col_1based_inclusive=annual_cashflow_col,
                description=f"Emily protected Annual Cashflow formula for {property_sheet_name}",
            )
        )

    _batch_sheet_requests(
        sheets=sheets,
        spreadsheet_id=spreadsheet_id,
        requests=requests,
        description="Format/protect linked formula cells",
    )


def _create_shannon_property_tab(
    sheets,
    spreadsheet_id: str,
    meta: dict,
    active_sheet_id: int,
    active_sheet_name: str,
    headers: dict[str, int],
    row_idx: int,
    address: str,
    city: str = "",
    state: str = "",
) -> int:
    sheet_title = _unique_sheet_title(meta, address)

    print(f"Creating native Google Sheets Shannon tab for {address}")

    shannon_output_path = _run_shannon_for_property(address, city=city, state=state)
    shannon_values = _shannon_workbook_to_values(shannon_output_path)

    property_sheet_id = _add_sheet(
        sheets,
        spreadsheet_id,
        sheet_title,
    )

    _write_values(
        sheets,
        spreadsheet_id,
        f"{_quote_sheet_name(sheet_title)}!A1",
        shannon_values,
    )

    _format_and_hide_property_tab_columns(
        sheets=sheets,
        spreadsheet_id=spreadsheet_id,
        property_sheet_id=property_sheet_id,
        shannon_values=shannon_values,
    )

    _link_property_tab_formulas(
        sheets=sheets,
        spreadsheet_id=spreadsheet_id,
        active_sheet_id=active_sheet_id,
        property_sheet_id=property_sheet_id,
        active_sheet_name=active_sheet_name,
        property_sheet_name=sheet_title,
        headers=headers,
        active_row_idx=row_idx,
    )

    return property_sheet_id


def _load_active_deals_state():
    _drive, sheets = _connect_services()
    spreadsheet_id, file_name = find_native_active_deals_sheet()

    meta = _get_spreadsheet_meta(sheets, spreadsheet_id)
    active_props = _get_active_sheet_props(meta)
    active_sheet_name = active_props["title"]
    active_sheet_id = active_props["sheetId"]

    values = _read_values(
        sheets,
        spreadsheet_id,
        f"{_quote_sheet_name(active_sheet_name)}!A:AC",
    )

    header_row, headers = _find_header_row_and_map(values)

    return sheets, spreadsheet_id, file_name, meta, active_sheet_name, active_sheet_id, values, header_row, headers


def _update_one_active_deal_block(body_text: str) -> ActiveDealsResult:
    sheets, spreadsheet_id, file_name, meta, active_sheet_name, active_sheet_id, values, header_row, headers = _load_active_deals_state()

    fields = _extract_labeled_fields(body_text)
    address = _extract_address_from_body(body_text, fields)

    if not address:
        raise ValueError("Active Deals email did not contain a usable address")

    match_key = normalize_address_match_key(address)

    if not match_key:
        raise ValueError(f"Could not build address match key for: {address}")

    address_col = _find_col(headers, ["Address", "Property Address", "Property", "Street Address"])

    if address_col is None:
        raise ValueError("Active Deals tab does not have an Address column")

    row_idx = _existing_match_row(values, header_row, address_col, match_key)
    action = "updated"

    if row_idx is None:
        row_idx = _first_empty_address_row(values, header_row, address_col)
        action = "added"

    city, state, _zip_code = _split_city_state_zip(address)
    updates: list[tuple[int, Any, str]] = []

    updates.append((address_col, address, "Address"))

    city_col = _find_col(headers, ["City"])
    state_col = _find_col(headers, ["St", "State"])

    if city and city_col:
        updates.append((city_col, city, "City"))

    if state and state_col:
        updates.append((state_col, state, "State"))

    for header_key, col_idx in headers.items():
        if col_idx == address_col:
            continue

        aliases = _field_aliases_for_header(header_key)
        value = None

        for alias in aliases:
            if _norm(alias) in fields:
                value = fields[_norm(alias)]
                break

        if value is None or value == "":
            continue

        raw_header = values[header_row - 1][col_idx - 1] if len(values[header_row - 1]) >= col_idx else header_key
        updates.append((col_idx, value, str(raw_header)))

    updated_fields = []

    for col_idx, value, field_name in updates:
        _write_values(
            sheets,
            spreadsheet_id,
            _a1(active_sheet_name, row_idx, col_idx),
            [[value]],
        )
        updated_fields.append(field_name)

    # After row update/add, ensure property tab exists and formulas are linked.
    meta = _get_spreadsheet_meta(sheets, spreadsheet_id)

    if not _sheet_exists_for_address(meta, address):
        # The Active Deals row is authoritative for lookup location. Re-read the
        # row after updates so City/State reach Shannon before tax/appraisal calls.
        current_row_values = _read_values(
            sheets,
            spreadsheet_id,
            f"{_quote_sheet_name(active_sheet_name)}!{row_idx}:{row_idx}",
        )
        current_row = current_row_values[0] if current_row_values else []
        city_for_shannon = (
            current_row[city_col - 1]
            if city_col and len(current_row) >= city_col
            else ""
        )
        state_for_shannon = (
            current_row[state_col - 1]
            if state_col and len(current_row) >= state_col
            else ""
        )

        _create_shannon_property_tab(
            sheets=sheets,
            spreadsheet_id=spreadsheet_id,
            meta=meta,
            active_sheet_id=active_sheet_id,
            active_sheet_name=active_sheet_name,
            headers=headers,
            row_idx=row_idx,
            address=address,
            city=str(city_for_shannon or "").strip(),
            state=str(state_for_shannon or "").strip() or "KS",
        )

    return ActiveDealsResult(
        action=action,
        file_name=file_name,
        row_number=row_idx,
        address=address,
        match_key=match_key,
        updated_fields=sorted(set(updated_fields)),
    )


def update_active_deals_from_email(body_text: str) -> list[ActiveDealsResult]:
    blocks = parse_active_deal_blocks(body_text)

    if not blocks:
        blocks = [body_text]

    results = []

    for block in blocks:
        result = _update_one_active_deal_block(block)
        results.append(result)

    return results


def ensure_active_deals_tabs_only() -> int:
    sheets, spreadsheet_id, _file_name, meta, active_sheet_name, active_sheet_id, values, header_row, headers = _load_active_deals_state()

    address_col = _find_col(headers, ["Address", "Property Address", "Property", "Street Address"])
    city_col = _find_col(headers, ["City"])
    state_col = _find_col(headers, ["St", "State"])

    if address_col is None:
        raise ValueError("Active Deals tab does not have an Address column")

    created_count = 0

    for row_idx in range(header_row + 1, len(values) + 1):
        row = values[row_idx - 1]
        address = row[address_col - 1] if len(row) >= address_col else ""
        address = str(address or "").strip()
        city = row[city_col - 1] if city_col and len(row) >= city_col else ""
        city = str(city or "").strip()
        state = row[state_col - 1] if state_col and len(row) >= state_col else ""
        state = str(state or "").strip() or "KS"

        if not address:
            continue

        meta = _get_spreadsheet_meta(sheets, spreadsheet_id)

        if _sheet_exists_for_address(meta, address):
            continue

        _create_shannon_property_tab(
            sheets=sheets,
            spreadsheet_id=spreadsheet_id,
            meta=meta,
            active_sheet_id=active_sheet_id,
            active_sheet_name=active_sheet_name,
            headers=headers,
            row_idx=row_idx,
            address=address,
            city=city,
            state=state,
        )

        created_count += 1

    return created_count
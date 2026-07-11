from __future__ import annotations

import argparse
import base64
import datetime as dt
import json
import re
import csv
import subprocess
import time
from pathlib import Path
from openpyxl import load_workbook
from app.morgan import handle_morgan_message
from app.active_deals_sheets import update_active_deals_from_email, ensure_active_deals_tabs_only
from html import unescape

GOOGLE_DRIVE_REMOTE_DIR = "gdrive:BLU Review Docs/Property_Reviews/Shannon_Output"
GOOGLE_DRIVE_ROOT_FOLDER_ID = "1kbUI4CrAXDMeE4mjj8pMJ_GrM488QEIa"

from app.config import load_settings
from app.gmail_client import GmailClient, get_subject, get_sender
from app.processor import (
    save_attachments,
    get_historian_request,
)
from app.state_store import StateStore
from app.address_organizer import organize_address_body_to_csv


PROPERTY_STATE_PATH = Path("/home/brubot77/.openclaw/workspace/shannon/property_state.json")
MONTHLY_DIR = "/home/brubot77/Monthly-Analyzer/input"
DEAL_DIR = "/home/brubot77/.openclaw/workspace/shannon/Input"
DEAL_OUTPUT_DIR = "/home/brubot77/.openclaw/workspace/shannon/Output"
MONTHLY_OUTPUT_DIR = Path("/home/brubot77/Monthly-Analyzer/output")

def handle_morgan_request(
    message: dict,
    gmail: GmailClient,
    token_path: str,
    processed_label_id: str,
    failed_label_id: str,
) -> bool:
    handled, success, detail = handle_morgan_message(
        message=message,
        gmail=gmail,
        token_path=token_path,
    )

    if not handled:
        return False

    message_id = message["id"]
    if not success:
        print(f"{message_id}: Morgan failed -> {detail}")
        gmail.mark_failed(message_id, failed_label_id)
        return True

    print(f"{message_id}: Morgan processed -> {detail}")
    gmail.mark_processed_and_archive(message_id, processed_label_id)
    return True

def trigger_deal_analyzer():
    cmd = (
        "cd /home/brubot77/.openclaw/workspace/shannon "
        "&& source .venv/bin/activate "
        "&& python3 -m shannon.cli"
    )

    result = subprocess.run(
        ["bash", "-lc", cmd],
        capture_output=True,
        text=True,
    )

    print(result.stdout)
    print(result.stderr)

    return result.returncode


def newest_output_after(before_snapshot):
    output_dir = Path(DEAL_OUTPUT_DIR)

    after_files = sorted(
        output_dir.glob("*.xlsx"),
        key=lambda p: p.stat().st_mtime,
        reverse=True,
    )

    for file in after_files:
        old_time = before_snapshot.get(str(file))
        if old_time is None:
            return file
        if file.stat().st_mtime > old_time:
            return file

    return None

def normalize_address_for_filename(address: str) -> str:
    """
    Convert first output address into a safe file-name prefix.

    Examples:
      1317 N Madison St, Wichita, KS -> 1317_N_Madison
      1317 N Madison Ave. -> 1317_N_Madison
    """

    text = str(address or "").strip()

    # Keep only street part before city/state.
    if "," in text:
        text = text.split(",", 1)[0].strip()

    text = text.replace("\u00a0", " ")

    # Remove punctuation.
    text = re.sub(r"[^A-Za-z0-9\s]", " ", text)

    # Remove city/state/zip noise if accidentally included.
    text = re.sub(r"\bwichita\b", " ", text, flags=re.IGNORECASE)
    text = re.sub(r"\bks\b|\bkansas\b", " ", text, flags=re.IGNORECASE)
    text = re.sub(r"\b\d{5}(?:-\d{4})?\b", " ", text)

    direction_replacements = {
        "north": "N",
        "south": "S",
        "east": "E",
        "west": "W",
    }

    street_suffixes = {
        "street", "st",
        "avenue", "ave",
        "road", "rd",
        "drive", "dr",
        "lane", "ln",
        "court", "ct",
        "place", "pl",
        "boulevard", "blvd",
        "terrace", "ter",
        "parkway", "pkwy",
        "circle", "cir",
        "trail", "trl",
        "way",
    }

    words = []

    for word in text.split():
        lower_word = word.lower().strip()

        if lower_word in street_suffixes:
            continue

        if lower_word in direction_replacements:
            words.append(direction_replacements[lower_word])
        else:
            words.append(word)

    safe = "_".join(words)
    safe = re.sub(r"_+", "_", safe).strip("_")

    return safe or "Shannon_Output"


def first_address_from_output_workbook(file_path: Path) -> str:
    """
    Read the first Address value from the Shannon output workbook.
    Shannon writes headers around row 15, but this searches flexibly.
    """

    wb = load_workbook(file_path, read_only=True, data_only=True)

    try:
        ws = wb.active

        address_col = None
        header_row = None

        # Search first 40 rows for Address header.
        for row in ws.iter_rows(min_row=1, max_row=40):
            for cell in row:
                if str(cell.value or "").strip().lower() == "address":
                    address_col = cell.column
                    header_row = cell.row
                    break

            if address_col and header_row:
                break

        if not address_col or not header_row:
            return ""

        # First non-empty Address value after the header.
        for row_num in range(header_row + 1, ws.max_row + 1):
            value = ws.cell(row=row_num, column=address_col).value
            value = str(value or "").strip()

            if value:
                return value

        return ""

    finally:
        wb.close()


def drive_filename_for_shannon_output(file_path: Path) -> str:
    first_address = first_address_from_output_workbook(file_path)
    address_prefix = normalize_address_for_filename(first_address)

    stamp = dt.datetime.now().strftime("%Y%m%d_%H%M%S")

    return f"{address_prefix}_{stamp}.xlsx"

def save_file_to_google_drive(file_path: Path) -> bool:
    drive_file_name = drive_filename_for_shannon_output(file_path)

    remote_path = (
        f"{GOOGLE_DRIVE_REMOTE_DIR.rstrip('/')}/"
        f"{drive_file_name}"
    )

    cmd = [
        "rclone",
        "copyto",
        str(file_path),
        remote_path,
        "--drive-root-folder-id",
        GOOGLE_DRIVE_ROOT_FOLDER_ID,
    ]

    result = subprocess.run(
        cmd,
        capture_output=True,
        text=True,
    )

    if result.stdout:
        print(result.stdout)

    if result.stderr:
        print(result.stderr)

    if result.returncode != 0:
        print(f"Google Drive upload failed for {file_path}")
        return False

    print(f"Saved Shannon output to Google Drive: {remote_path}")
    return True


def canonical_property_key(
    address: str | None,
    city: str | None = "",
    state: str | None = "",
    zip_code: str | None = "",
) -> str:
    """
    Loose street-only canonical key for property_state.json.

    City/state/ZIP intentionally do not matter.
    Directionals and street suffixes are removed so these all match:
      2026 N Arkansas Ave
      2026 Arkansas
      2026 Arkansas St
      2026 S Arkansas Blvd

    Result:
      2026 arkansas
    """

    text = str(address or "").strip().lower()

    # If city/state are included after commas, keep only street portion.
    if "," in text:
        text = text.split(",", 1)[0]

    # Cut off accidentally swallowed labeled fields.
    text = re.split(
        r"\bstatus\s*:|\bzest rent\s*:|\bnotes\s*:",
        text,
        flags=re.IGNORECASE,
    )[0]

    # Remove city/state/ZIP noise if typed without commas.
    text = re.sub(r"\bwichita\b", " ", text, flags=re.IGNORECASE)
    text = re.sub(r"\bks\b|\bkansas\b", " ", text, flags=re.IGNORECASE)
    text = re.sub(r"\b\d{5}(?:-\d{4})?\b", " ", text)

    # Remove unit/apartment info.
    text = re.sub(r"\b(apartment|apt|unit|ste|suite)\s+\w+\b", " ", text)

    # Remove punctuation.
    text = re.sub(r"[^a-z0-9\s]", " ", text)

    direction_words = {
        "n", "s", "e", "w",
        "ne", "nw", "se", "sw",
        "north", "south", "east", "west",
        "northeast", "northwest", "southeast", "southwest",
    }

    street_suffixes = {
        "street", "st",
        "avenue", "ave", "av",
        "road", "rd",
        "drive", "dr",
        "lane", "ln",
        "court", "ct",
        "place", "pl",
        "boulevard", "blvd",
        "terrace", "ter",
        "parkway", "pkwy",
        "circle", "cir",
        "trail", "trl",
        "highway", "hwy",
        "way",
    }

    words = []

    for word in text.split():
        if word in direction_words:
            continue

        if word in street_suffixes:
            continue

        words.append(word)

    text = " ".join(words)
    text = re.sub(r"\s+", " ", text).strip()

    return text


def find_existing_property_state_key(
    property_state: dict,
    property_key: str,
) -> str | None:
    """
    Finds an existing property_state key using the new loose address rules.

    This prevents duplicate records when old saved keys still contain a
    direction, such as "2026 n arkansas", but the new update email says
    "2026 Arkansas".
    """

    if property_key in property_state:
        return property_key

    for existing_key, state_entry in property_state.items():
        if canonical_property_key(existing_key) == property_key:
            return existing_key

        if isinstance(state_entry, dict):
            display_address = state_entry.get("display_address", "")

            if canonical_property_key(display_address) == property_key:
                return existing_key

    return None

def load_property_state() -> dict:
    if not PROPERTY_STATE_PATH.exists():
        return {}

    try:
        return json.loads(PROPERTY_STATE_PATH.read_text(encoding="utf-8"))
    except Exception:
        return {}


def save_property_state(state: dict) -> None:
    PROPERTY_STATE_PATH.write_text(
        json.dumps(state, indent=2),
        encoding="utf-8",
    )


def extract_reply_to_from_body(body_text: str) -> str | None:
    patterns = [
        r"^\s*received\s+from\s*:\s*([^\s,;<>]+@[^\s,;<>]+)",
        r"^\s*to\s*:\s*([^\s,;<>]+@[^\s,;<>]+)",
        r"^\s*send\s+to\s*:\s*([^\s,;<>]+@[^\s,;<>]+)",
        r"^\s*email\s*:\s*([^\s,;<>]+@[^\s,;<>]+)",
    ]

    for line in (body_text or "").splitlines():
        for pattern in patterns:
            match = re.search(pattern, line, re.IGNORECASE)
            if match:
                return match.group(1).strip()

    return None


def decode_message_body(message: dict) -> str:
    payload = message.get("payload", {})
    plain_texts: list[str] = []
    html_texts: list[str] = []

    def decode_data(data: str) -> str:
        padded = data + "=" * (-len(data) % 4)
        return base64.urlsafe_b64decode(
            padded.encode("utf-8")
        ).decode("utf-8", errors="ignore")

    def html_to_text(html: str) -> str:
        text = html

        text = re.sub(r"(?is)<(script|style).*?>.*?</\1>", " ", text)
        text = re.sub(r"(?i)<br\s*/?>", "\n", text)
        text = re.sub(r"(?i)</p\s*>", "\n", text)
        text = re.sub(r"(?i)</div\s*>", "\n", text)
        text = re.sub(r"(?i)</li\s*>", "\n", text)
        text = re.sub(r"(?s)<[^>]+>", " ", text)
        text = unescape(text)
        text = text.replace("\xa0", " ")
        text = re.sub(r"[ \t]+", " ", text)
        text = re.sub(r"\n\s+", "\n", text)
        text = re.sub(r"\n{3,}", "\n\n", text)

        return text.strip()

    def walk(part: dict) -> None:
        mime_type = part.get("mimeType", "")
        body = part.get("body", {})
        data = body.get("data")

        if data:
            try:
                decoded = decode_data(data)

                if mime_type == "text/plain" and decoded.strip():
                    plain_texts.append(decoded.strip())

                elif mime_type == "text/html" and decoded.strip():
                    converted = html_to_text(decoded)

                    if converted.strip():
                        html_texts.append(converted.strip())

            except Exception:
                pass

        for child in part.get("parts", []) or []:
            walk(child)

    walk(payload)

    if plain_texts:
        return "\n".join(plain_texts).strip()

    if html_texts:
        return "\n".join(html_texts).strip()

    return ""

def parse_address_update_body(body_text: str) -> list[dict[str, str]]:
    """
    Parses one or more Update Address blocks.

    Required:
      Address:

    Optional:
      Status:
      Zest Rent:
      Notes:

    City/state are not required and are ignored for matching.
    """

    text = (body_text or "").replace("\r\n", "\n").replace("\r", "\n").strip()

    if not text:
        return []

    label_re = re.compile(
        r"(?im)^(address|status update|status|zest rent|notes)\s*:\s*"
    )

    matches = list(label_re.finditer(text))

    if not matches:
        return []

    blocks: list[dict[str, str]] = []
    current: dict[str, str] | None = None

    for i, match in enumerate(matches):
        label = match.group(1).lower().strip()
        start = match.end()
        end = matches[i + 1].start() if i + 1 < len(matches) else len(text)

        value = text[start:end].strip()
        value = re.sub(r"\n+", " ", value)
        value = re.sub(r"\s+", " ", value).strip()

        if label == "address":
            if current and current.get("address"):
                blocks.append(current)

            current = {
                "address": value,
                "status": "",
                "zest_rent": "",
                "notes": "",
            }

            continue

        if current is None:
            continue

        if label in {"status", "status update"}:
            current["status"] = value
        elif label == "zest rent":
            current["zest_rent"] = value
        elif label == "notes":
            current["notes"] = value

    if current and current.get("address"):
        blocks.append(current)

    return blocks


def handle_address_update_request(
    message: dict,
    sender: str,
    gmail: GmailClient,
    processed_label_id: str,
    failed_label_id: str,
) -> bool:
    subject = get_subject(message).strip().lower()

    if subject != "update address":
        return False

    message_id = message["id"]
    body_text = decode_message_body(message)
    parsed_updates = parse_address_update_body(body_text)

    if not parsed_updates:
        gmail.mark_failed(message_id, failed_label_id)
        print(f"{message_id}: update address email contained no valid update blocks")
        return True

    now = dt.datetime.now(dt.UTC).isoformat()
    property_state = load_property_state()

    updated_count = 0

    for parsed in parsed_updates:
        address = parsed.get("address", "")
        new_status = parsed.get("status", "")
        new_zest_rent = parsed.get("zest_rent", "")
        new_note = parsed.get("notes", "")

        if not address:
            print(f"{message_id}: skipping update block with missing address")
            continue

        property_key = canonical_property_key(address)
        existing_property_key = find_existing_property_state_key(
            property_state,
            property_key,
        )

        if existing_property_key:
            state_entry = property_state[existing_property_key]

            # Migrate old saved keys, such as "2026 n arkansas", to the
            # new loose key, such as "2026 arkansas".
            if existing_property_key != property_key:
                property_state[property_key] = property_state.pop(existing_property_key)
                state_entry = property_state[property_key]
                print(
                    f"{message_id}: migrated property key "
                    f"{existing_property_key} -> {property_key}"
                )
        else:
            state_entry = {
                "display_address": address,
                "status": "Under Review",
                "zest_rent": "",
                "notes_history": [],
                "first_seen_utc": now,
                "last_seen_utc": now,
            }

            property_state[property_key] = state_entry

        state_entry["display_address"] = address.split(",", 1)[0].strip()
        state_entry["last_seen_utc"] = now
        state_entry.setdefault("status", "Under Review")
        state_entry.setdefault("zest_rent", "")
        state_entry.setdefault("notes_history", [])

        if new_status:
            state_entry["status"] = new_status

        if new_zest_rent:
            state_entry["zest_rent"] = new_zest_rent

        if new_note:
            state_entry["notes_history"].append(
                {
                    "timestamp_utc": now,
                    "sender": sender,
                    "note": new_note,
                }
            )

        updated_count += 1

    if updated_count == 0:
        gmail.mark_failed(message_id, failed_label_id)
        print(f"{message_id}: no usable update blocks found in Update Address email")
        return True

    save_property_state(property_state)

    gmail.mark_processed_and_archive(
        message_id,
        processed_label_id,
    )

    print(f"{message_id}: updated {updated_count} property record(s) from Update Address email")

    return True


def handle_historian_request(
    message: dict,
    gmail: GmailClient,
    processed_label_id: str,
    failed_label_id: str,
) -> bool:
    subject = get_subject(message)
    historian_file = get_historian_request(subject)

    if not historian_file:
        return False

    message_id = message["id"]

    historian_path = MONTHLY_OUTPUT_DIR / historian_file

    if not historian_path.exists():
        print(f"{message_id}: historian file missing -> {historian_path}")

        gmail.mark_failed(
            message_id,
            failed_label_id,
        )

        return True

    print(f"{message_id}: historian retrieval matched -> {historian_path}")

    gmail.reply_with_attachment(
        original_message=message,
        attachment_path=str(historian_path),
        body_text=f"Attached is the requested historian file: {historian_file}",
    )

    if historian_file in {"BLU1_historian.xlsx", "BLU2_historian.xlsx"}:
        consolidation_path = MONTHLY_OUTPUT_DIR / "BLU_consolidation.xlsx"

        if consolidation_path.exists():
            print(f"{message_id}: sending BLU consolidation workbook -> {consolidation_path}")

            gmail.reply_with_attachment(
                original_message=message,
                attachment_path=str(consolidation_path),
                body_text="Attached is the current BLU consolidation workbook.",
            )

        else:
            print(f"{message_id}: BLU consolidation workbook missing -> {consolidation_path}")

    gmail.mark_processed_and_archive(
        message_id,
        processed_label_id,
    )

    return True


def handle_address_data_request(
    message: dict,
    gmail: GmailClient,
    processed_label_id: str,
    failed_label_id: str,
) -> bool:
    subject = get_subject(message).strip().lower()

    if subject != "address data":
        return False

    message_id = message["id"]
    body_text = decode_message_body(message)

    if not body_text.strip():
        print(f"{message_id}: Address Data email had no body text")
        gmail.mark_failed(message_id, failed_label_id)
        return True

    csv_path = organize_address_body_to_csv(body_text)
    
    with open(csv_path, "r", encoding="utf-8-sig", newline="") as f:
        rows = list(csv.DictReader(f))

    if not rows:
        print(f"{message_id}: Address Data created CSV but found no address rows")
        gmail.mark_failed(message_id, failed_label_id)
        return True 
    
    original_sender = get_sender(message)

    print(f"{message_id}: Address Organizer created CSV -> {csv_path}")
    print(
        f"{message_id}: forwarding run shannon CSV to bru.bot77@gmail.com "
        f"with received from={original_sender}"
    )

    gmail.reply_with_attachment(
        original_message=message,
        attachment_path=str(csv_path),
        body_text=f"received from: {original_sender}",
        subject="run shannon",
        to_override="bru.bot77@gmail.com",
    )

    gmail.mark_processed_and_archive(
        message_id,
        processed_label_id,
    )

    return True

def handle_active_deals_request(
    message: dict,
    gmail: GmailClient,
    processed_label_id: str,
    failed_label_id: str,
) -> bool:
    subject = get_subject(message).strip().lower()

    if "active deals" not in subject:
        return False

    message_id = message["id"]
    body_text = decode_message_body(message)

    # Subject "active deals" with no real update data should run the tab-check-only flow.
    # Some blank emails decode as non-empty because Gmail may include hidden HTML,
    # quoted text, signatures, or formatting artifacts.
    body_for_update = (body_text or "").strip()

    has_active_deal_update_data = bool(
        re.search(
            r"(?im)^\s*(address|status|status update|zest rent|notes|latest offer|rehab est|appraisal est)\s*:",
            body_for_update,
        )
        or re.search(
            r"(?im)^\s*\d{2,6}\s+(?:[NSEW]\.?\s+|North\s+|South\s+|East\s+|West\s+)?[A-Za-z0-9.'-]+",
            body_for_update,
        )
    )

    if not has_active_deal_update_data:
        try:
            created_tab_count = ensure_active_deals_tabs_only()
        except Exception as exc:
            print(f"{message_id}: Active Deals tab check failed -> {exc}")
            return False

        print(
            f"{message_id}: Active Deals email had no usable update body; "
            f"created {created_tab_count} missing property underwriting tab(s)"
        )
        return True

    try:
        results = update_active_deals_from_email(body_text)
    except Exception as exc:
        print(f"{message_id}: Active Deals update failed -> {exc}")
        return False

    for result in results:
        print(
            f"{message_id}: Active Deals {result.action} row {result.row_number} "
            f"in {result.file_name} for {result.address} "
            f"match_key={result.match_key} "
            f"fields={result.updated_fields}"
        )

    gmail.mark_processed_and_archive(
        message_id,
        processed_label_id,
    )

    return True

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--process", action="store_true")
    args = parser.parse_args()

    settings = load_settings()

    state = StateStore(settings.state_file)
    processed_ids = state.load()

    gmail = GmailClient(
        settings.gmail_credentials_path,
        settings.gmail_token_path,
    )

    message_ids = gmail.list_message_ids(settings.gmail_query)

    processed_label_id = gmail.create_label_if_missing(
        settings.processed_label
    )

    failed_label_id = gmail.create_label_if_missing(
        settings.failed_label
    )

    deal_requests = {}

    print(f"Found {len(message_ids)} matching message(s)")

    for message_id in message_ids:
        if message_id in processed_ids:
            continue

        message = gmail.get_message(message_id)

        handled_morgan = handle_morgan_request(
            message,
            gmail,
            settings.gmail_token_path,
            processed_label_id,
            failed_label_id,
        )

        if handled_morgan:
            processed_ids.add(message_id)
            state.save(processed_ids)
            continue

        handled_address_data = handle_address_data_request(
            message,
            gmail,
            processed_label_id,
            failed_label_id,
        )

        if handled_address_data:
            processed_ids.add(message_id)
            state.save(processed_ids)
            continue

        handled_active_deals = handle_active_deals_request(
            message,
            gmail,
            processed_label_id,
            failed_label_id,
        )

        if handled_active_deals:
            processed_ids.add(message_id)
            state.save(processed_ids)
            continue

        handled_historian = handle_historian_request(
            message,
            gmail,
            processed_label_id,
            failed_label_id,
        )

        if handled_historian:
            processed_ids.add(message_id)
            state.save(processed_ids)
            continue

        sender = get_sender(message)

        handled_update = handle_address_update_request(
            message,
            sender,
            gmail,
            processed_label_id,
            failed_label_id,
        )

        if handled_update:
            processed_ids.add(message_id)
            state.save(processed_ids)
            continue

        saved_paths = save_attachments(
            message,
            gmail,
            settings.monthly_input_dir,
            settings.deal_input_dir,
            settings.unmatched_dir,
        )

        for path in saved_paths:
            if path.startswith(DEAL_DIR):
                print(f"{message_id}: CSV saved for Shannon run -> {path}")
                deal_requests[message_id] = {
                    "message": message,
                    "csv_path": path,
                }

        if saved_paths and message_id not in deal_requests:
            gmail.mark_processed_and_archive(
                message_id,
                processed_label_id,
            )

            processed_ids.add(message_id)
            state.save(processed_ids)

            print(f"{message_id}: saved attachment(s), marked processed")
    if deal_requests:
        print("Triggering Deal Analyzer")

        output_dir = Path(DEAL_OUTPUT_DIR)

        before_snapshot = {
            str(p): p.stat().st_mtime
            for p in output_dir.glob("*.xlsx")
        }

        rc = trigger_deal_analyzer()

        if rc != 0:
            print("Deal Analyzer failed")

            for message_id in deal_requests:
                gmail.mark_failed(message_id, failed_label_id)

            return

        time.sleep(2)

        new_file = newest_output_after(before_snapshot)

        if not new_file:
            print("No new Excel output detected")

            for message_id in deal_requests:
                gmail.mark_failed(message_id, failed_label_id)

            return

        print(f"New Shannon output detected: {new_file}")

        for message_id, request in deal_requests.items():
            message = request["message"]
            csv_path = Path(request["csv_path"])

            body_text = decode_message_body(message)

            to_addr = extract_reply_to_from_body(body_text)

            if to_addr:
                print(f"{message_id}: sending Shannon results to body recipient -> {to_addr}")
            else:
                sender_email = get_sender(message)
                to_addr = sender_email

                print(
                    f"{message_id}: no 'received from' email found "
                    f"-> falling back to sender {sender_email}"
                )

            run_date = dt.datetime.now().strftime("%d-%b-%Y")
            csv_file_name = csv_path.name

            outbound_subject = (
                f"{csv_file_name} {run_date} "
                f"prelim underwriting data ready for review"
            )

            uploaded = save_file_to_google_drive(new_file)

            if not uploaded:
                gmail.mark_failed(message_id, failed_label_id)
                print(f"{message_id}: failed to save Shannon output to Google Drive")
                continue

            print(f"{message_id}: Shannon output saved to Google Drive -> {new_file.name}")

            gmail.mark_processed_and_archive(
                message_id,
                processed_label_id,
            )

            processed_ids.add(message_id)

        state.save(processed_ids)

    return


if __name__ == "__main__":
    main()
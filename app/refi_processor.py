from __future__ import annotations

import base64
import datetime as dt
import re
import shutil
import tempfile
from pathlib import Path

from openpyxl import Workbook, load_workbook
from pypdf import PdfReader


REFI_ROOT = Path("/home/brubot77/Refi")
REFI_BLU_DIR = REFI_ROOT / "BLU"
REFI_BLU2_DIR = REFI_ROOT / "BLU2"
REFI_LOG_PATH = REFI_ROOT / "refi_log.xlsx"

REFI_LOG_HEADERS = [
    "Received UTC",
    "LLC",
    "Property Address",
    "Stored File",
    "Original Filename",
    "Lender",
    "Settlement Date",
    "New Loan Amount",
    "Loan Origination Fee",
    "Loan Document Fee",
    "Appraisal Fee",
    "Flood Certification",
    "Loan Policy",
    "Loan Policy Credit",
    "Closing Fee",
    "Overnight Delivery Fee",
    "Record Mortgage",
    "Record Assignment of Rents",
    "SubTotal Debits",
    "SubTotal Credits",
    "Due to Borrower",
    "Escrow/File No",
    "Message ID",
    "Sender",
]


def ensure_refi_folders() -> None:
    REFI_BLU_DIR.mkdir(parents=True, exist_ok=True)
    REFI_BLU2_DIR.mkdir(parents=True, exist_ok=True)


def sanitize_filename(value: str) -> str:
    value = str(value or "").strip()
    value = re.sub(r"[^A-Za-z0-9._ -]+", "_", value)
    value = re.sub(r"\s+", "_", value)
    value = value.strip("._- ")
    return value or "refi"


def iter_message_parts(payload: dict):
    yield payload

    for part in payload.get("parts", []) or []:
        yield from iter_message_parts(part)


def save_refi_pdf_attachments(message: dict, gmail) -> list[tuple[Path, str]]:
    """
    Save PDF attachments from a Gmail message to a temp folder.
    """

    saved: list[tuple[Path, str]] = []
    message_id = message["id"]
    payload = message.get("payload", {})

    temp_dir = Path(tempfile.gettempdir()) / "emily_refi"
    temp_dir.mkdir(parents=True, exist_ok=True)

    for part in iter_message_parts(payload):
        filename = part.get("filename") or ""

        if not filename.lower().endswith(".pdf"):
            continue

        body = part.get("body", {}) or {}
        data = body.get("data")
        attachment_id = body.get("attachmentId")

        if attachment_id:
            attachment = (
                gmail.service.users()
                .messages()
                .attachments()
                .get(
                    userId="me",
                    messageId=message_id,
                    id=attachment_id,
                )
                .execute()
            )
            data = attachment.get("data")

        if not data:
            continue

        padded = data + "=" * (-len(data) % 4)
        content = base64.urlsafe_b64decode(padded.encode("utf-8"))

        safe_name = sanitize_filename(filename)
        output_path = temp_dir / f"{message_id}_{safe_name}"
        output_path.write_bytes(content)
        saved.append((output_path, filename))

    return saved


def extract_pdf_text(pdf_path: Path) -> str:
    reader = PdfReader(str(pdf_path))
    pages = []

    for page in reader.pages:
        try:
            pages.append(page.extract_text() or "")
        except Exception:
            continue

    return "\n".join(pages)


def extract_line_value(text: str, label: str) -> str:
    pattern = rf"{re.escape(label)}\s*:\s*(.+)"
    match = re.search(pattern, text, re.IGNORECASE)

    if not match:
        return ""

    return match.group(1).strip()


def money_after_label(text: str, label: str) -> str:
    pattern = rf"{re.escape(label)}.*?(\$[\d,]+\.\d{{2}})"
    match = re.search(pattern, text, re.IGNORECASE)

    return match.group(1) if match else ""

def normalize_money(value: str) -> str:
    value = str(value or "").strip()
    match = re.search(r"\$[\d,]+\.\d{2}", value)
    return match.group(0) if match else ""


def extract_all_money_values(text: str) -> dict[str, str]:
    """
    Extracts refi dollar values from ALTA-style settlement statements.
    """

    values = {
        "new_loan_amount": money_after_label(text, "New Loan Amount"),
        "loan_origination_fee": money_after_label(text, "Loan Origination Fee"),
        "loan_document_fee": money_after_label(text, "Loan Document Fee"),
        "appraisal_fee": money_after_label(text, "Appraisal Fee"),
        "flood_certification": money_after_label(text, "Flood Certification"),
        "loan_policy": "",
        "loan_policy_credit": money_after_label(text, "Loan Policy Credit"),
        "closing_fee": money_after_label(text, "Closing Fee"),
        "overnight_delivery_fee": money_after_label(text, "Overnight Delivery Fee"),
        "record_mortgage": money_after_label(text, "Record Mortgage"),
        "record_assignment_of_rents": money_after_label(text, "Record Assignment of Rents"),
        "subtotal_debits": "",
        "subtotal_credits": "",
        "due_to_borrower": money_after_label(text, "Due to Buyer/Borrower"),
    }

    # Loan Policy line often includes the policy amount first and the fee second:
    # "Loan Policy $385,600.00 $1,618.00"
    loan_policy_match = re.search(
        r"Loan Policy\s+\$[\d,]+\.\d{2}\s+(\$[\d,]+\.\d{2})",
        text,
        re.IGNORECASE,
    )
    if loan_policy_match:
        values["loan_policy"] = loan_policy_match.group(1)
    else:
        values["loan_policy"] = money_after_label(text, "Loan Policy")

    # SubTotals line usually has debit and credit:
    # "SubTotals $8,991.00 $386,004.00"
    subtotal_match = re.search(
        r"SubTotals\s+(\$[\d,]+\.\d{2})\s+(\$[\d,]+\.\d{2})",
        text,
        re.IGNORECASE,
    )
    if subtotal_match:
        values["subtotal_debits"] = subtotal_match.group(1)
        values["subtotal_credits"] = subtotal_match.group(2)

    return values

def detect_llc(text: str) -> str:
    lower = text.lower()

    if "blu realty group 2 llc" in lower:
        return "BLU2"

    if "blu realty group llc" in lower:
        return "BLU"

    # fallback from filename/content
    if "blu2" in lower or "blu 2" in lower:
        return "BLU2"

    if "blu1" in lower or "blu 1" in lower or "blu" in lower:
        return "BLU"

    return "UNKNOWN"


def extract_property_addresses(text: str) -> list[str]:
    """
    Handles ALTA line like:
    Property Address: 711 Old Main St, 424 E 8th St, 319 W 8th St,
    1020 S Poplar St, Newton, KS 67114
    """

    match = re.search(
        r"Property Address\s*:\s*(.+?)(?:\nLegal\s*:|\nBuyer\s*:|\nSeller\s*:|\nLender\s*:)",
        text,
        re.IGNORECASE | re.DOTALL,
    )

    if not match:
        return []

    raw = match.group(1)
    raw = re.sub(r"\s+", " ", raw).strip()

    parts = [p.strip() for p in raw.split(",") if p.strip()]

    if len(parts) >= 3:
        city = parts[-2]
        state_zip = parts[-1]
        street_parts = parts[:-2]

        addresses = [
            f"{street}, {city}, {state_zip}"
            for street in street_parts
            if street
        ]

        return addresses

    return [raw]


def extract_refi_info(pdf_path: Path) -> dict:
    text = extract_pdf_text(pdf_path)

    escrow_no = ""
    escrow_match = re.search(
        r"File No\.?/Escrow No\.?\s*:\s*([^\n]+)",
        text,
        re.IGNORECASE,
    )

    if escrow_match:
        escrow_no = escrow_match.group(1).strip()

    llc = detect_llc(text)
    addresses = extract_property_addresses(text)

    money_values = extract_all_money_values(text)

    return {
        "llc": llc,
        "addresses": addresses,
        "lender": extract_line_value(text, "Lender"),
        "settlement_date": extract_line_value(text, "Settlement Date"),
        "escrow_no": escrow_no,
        "text": text,
        **money_values,
    }


def refi_destination_for_llc(llc: str) -> Path:
    if llc == "BLU2":
        return REFI_BLU2_DIR

    if llc == "BLU":
        return REFI_BLU_DIR

    return REFI_ROOT / "UNKNOWN"


def create_stored_refi_filename(info: dict, original_filename: str) -> str:
    date_part = dt.datetime.now().strftime("%Y%m%d_%H%M%S")
    llc = info.get("llc") or "UNKNOWN"

    addresses = info.get("addresses") or []
    first_address = addresses[0] if addresses else "refi"
    first_address = first_address.split(",", 1)[0]

    name = f"{llc}_{first_address}_{date_part}.pdf"
    return sanitize_filename(name)


def ensure_refi_log() -> None:
    REFI_ROOT.mkdir(parents=True, exist_ok=True)

    if REFI_LOG_PATH.exists():
        return

    wb = Workbook()
    ws = wb.active
    ws.title = "Refi Log"
    ws.append(REFI_LOG_HEADERS)
    wb.save(REFI_LOG_PATH)

def normalize_duplicate_key(value: str) -> str:
    value = str(value or "").strip().lower()
    value = re.sub(r"\s+", " ", value)
    return value


def existing_refi_log_keys(ws) -> set[tuple[str, str]]:
    """
    Returns existing duplicate keys:
      (Original Filename, Property Address)
    """

    keys: set[tuple[str, str]] = set()

    header_row = [cell.value for cell in ws[1]]
    header_map = {
        str(header or "").strip(): idx + 1
        for idx, header in enumerate(header_row)
    }

    original_col = header_map.get("Original Filename")
    address_col = header_map.get("Property Address")

    if not original_col or not address_col:
        return keys

    for row in range(2, ws.max_row + 1):
        original_filename = ws.cell(row=row, column=original_col).value
        property_address = ws.cell(row=row, column=address_col).value

        if not original_filename or not property_address:
            continue

        keys.add(
            (
                normalize_duplicate_key(original_filename),
                normalize_duplicate_key(property_address),
            )
        )

    return keys

def append_refi_log_rows(
    info: dict,
    stored_path: Path,
    original_filename: str,
    message_id: str,
    sender: str,
) -> int:
    """
    Append one row per property address.

    Avoid duplicate rows when the same:
      Original Filename + Property Address
    already exists.
    """

    ensure_refi_log()

    wb = load_workbook(REFI_LOG_PATH)
    ws = wb["Refi Log"]

    existing_keys = existing_refi_log_keys(ws)

    received_utc = dt.datetime.now(dt.UTC).isoformat()
    addresses = info.get("addresses") or []

    added_count = 0

    for address in addresses:
        duplicate_key = (
            normalize_duplicate_key(original_filename),
            normalize_duplicate_key(address),
        )

        if duplicate_key in existing_keys:
            continue

        ws.append(
            [
                received_utc,
                info.get("llc", ""),
                address,
                str(stored_path),
                original_filename,
                info.get("lender", ""),
                info.get("settlement_date", ""),
                info.get("new_loan_amount", ""),
                info.get("loan_origination_fee", ""),
                info.get("loan_document_fee", ""),
                info.get("appraisal_fee", ""),
                info.get("flood_certification", ""),
                info.get("loan_policy", ""),
                info.get("loan_policy_credit", ""),
                info.get("closing_fee", ""),
                info.get("overnight_delivery_fee", ""),
                info.get("record_mortgage", ""),
                info.get("record_assignment_of_rents", ""),
                info.get("subtotal_debits", ""),
                info.get("subtotal_credits", ""),
                info.get("due_to_borrower", ""),
                info.get("escrow_no", ""),
                message_id,
                sender,
            ]
        )

        existing_keys.add(duplicate_key)
        added_count += 1

    wb.save(REFI_LOG_PATH)
    return added_count


def process_refi_message(message: dict, gmail, sender: str) -> tuple[bool, str]:
    """
    Returns:
      (success, message)
    """

    ensure_refi_folders()

    message_id = message["id"]
    pdfs = save_refi_pdf_attachments(message, gmail)

    if not pdfs:
        return False, "No PDF attachments found"

    processed_count = 0

    total_rows_added = 0

    for pdf_path, original_filename in pdfs:
        info = extract_refi_info(pdf_path)

        if not info.get("addresses"):
            return False, f"No property addresses found in {pdf_path.name}"

        if info.get("llc") == "UNKNOWN":
            return False, f"Could not determine BLU vs BLU2 for {pdf_path.name}"

        destination_dir = refi_destination_for_llc(info["llc"])
        destination_dir.mkdir(parents=True, exist_ok=True)

        stored_name = create_stored_refi_filename(info, original_filename)
        stored_path = destination_dir / stored_name

        shutil.copy2(pdf_path, stored_path)

        rows_added = append_refi_log_rows(
            info=info,
            stored_path=stored_path,
            original_filename=original_filename,
            message_id=message_id,
            sender=sender,
        )

        total_rows_added += rows_added

        processed_count += 1

    return True, f"Processed {processed_count} refi PDF(s), added {total_rows_added} log row(s)"
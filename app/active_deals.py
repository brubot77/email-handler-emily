from __future__ import annotations

import os
import re
import tempfile
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any

from google.auth.transport.requests import Request
from google.oauth2.credentials import Credentials
from googleapiclient.discovery import build
from googleapiclient.http import MediaFileUpload, MediaIoBaseDownload
from openpyxl import load_workbook


DRIVE_SCOPE = "https://www.googleapis.com/auth/drive"

DEFAULT_CREDENTIALS_PATH = "/home/brubot77/email-handler-emily/credentials.json"
DEFAULT_TOKEN_PATH = "/home/brubot77/email-handler-emily/token.json"

ACTIVE_DEALS_NAME_RE = re.compile(
    r"^BLU Active Deals(?:\s*\(.*\)|\s*-\s*.*)?(?:\.xlsx)?$",
    re.IGNORECASE,
)


@dataclass
class ActiveDealsResult:
    action: str
    file_name: str
    row_number: int
    address: str
    match_key: str
    updated_fields: list[str]


def _norm(value: Any) -> str:
    return re.sub(r"[^a-z0-9]+", "", str(value or "").strip().lower())


def _clean_text(value: Any) -> str:
    return str(value or "").strip()


def normalize_address_match_key(address: str) -> str:
    """
    Match only street number + street name.

    Ignore:
      - direction: N, S, E, W, North, South, East, West
      - suffix: St, Ave, Rd, Dr, etc.
      - city/state/zip after commas
    """

    text = str(address or "").strip().lower()

    if "," in text:
        text = text.split(",", 1)[0]

    text = re.sub(r"\bwichita\b|\bks\b|\bkansas\b", " ", text, flags=re.IGNORECASE)
    text = re.sub(r"\b\d{5}(?:-\d{4})?\b", " ", text)

    text = re.sub(r"[^a-z0-9\s]", " ", text)

    directions = {
        "n", "s", "e", "w",
        "ne", "nw", "se", "sw",
        "north", "south", "east", "west",
        "northeast", "northwest", "southeast", "southwest",
    }

    suffixes = {
        "street", "st",
        "avenue", "ave",
        "road", "rd",
        "drive", "dr",
        "lane", "ln",
        "court", "ct",
        "place", "pl",
        "boulevard", "blvd",
        "terrace", "ter",
        "parkway", "pkwy",
        "circle", "cir",
        "trail", "trl",
        "way",
        "loop",
        "highway", "hwy",
        "route",
    }

    unit_words = {
        "apt", "apartment", "unit", "suite", "ste",
    }

    words = []
    skip_next = False

    for word in text.split():
        if skip_next:
            skip_next = False
            continue

        if word in unit_words:
            skip_next = True
            continue

        if word in directions:
            continue

        if word in suffixes:
            continue

        words.append(word)

    if not words:
        return ""

    number = next((w for w in words if w.isdigit()), "")

    if not number:
        return " ".join(words[:2]).strip()

    try:
        number_idx = words.index(number)
    except ValueError:
        number_idx = 0

    street_name = ""
    for word in words[number_idx + 1:]:
        if not word.isdigit():
            street_name = word
            break

    if street_name:
        return f"{number} {street_name}".strip()

    return number


def _connect_drive():
    credentials_path = Path(os.getenv("ACTIVE_DEALS_CREDENTIALS_PATH", DEFAULT_CREDENTIALS_PATH))
    token_path = Path(os.getenv("ACTIVE_DEALS_TOKEN_PATH", DEFAULT_TOKEN_PATH))

    if not credentials_path.exists():
        raise FileNotFoundError(f"Google credentials file not found: {credentials_path}")

    if not token_path.exists():
        raise FileNotFoundError(f"Google token file not found: {token_path}")

    creds = Credentials.from_authorized_user_file(
        str(token_path),
        scopes=[DRIVE_SCOPE],
    )

    if not creds.valid:
        if creds.expired and creds.refresh_token:
            creds.refresh(Request())
            token_path.write_text(creds.to_json(), encoding="utf-8")
        else:
            raise RuntimeError("Google token is invalid. Re-run OAuth with Drive scope.")

    return build("drive", "v3", credentials=creds)


def find_latest_active_deals_file() -> tuple[str, str]:
    service = _connect_drive()

    query = (
        "name contains 'BLU Active Deals' "
        "and mimeType != 'application/vnd.google-apps.folder' "
        "and trashed = false"
    )

    response = service.files().list(
        q=query,
        spaces="drive",
        fields="files(id,name,mimeType,modifiedTime)",
        includeItemsFromAllDrives=True,
        supportsAllDrives=True,
        pageSize=100,
    ).execute()

    candidates: list[tuple[str, str, str]] = []

    for item in response.get("files", []):
        name = item.get("name", "")

        if not ACTIVE_DEALS_NAME_RE.match(name):
            continue

        candidates.append((
            item.get("modifiedTime", ""),
            item["id"],
            name,
        ))

    if not candidates:
        raise FileNotFoundError("No Google Drive file found matching BLU Active Deals*.xlsx")

    candidates.sort(reverse=True)
    _modified_time, file_id, name = candidates[0]

    return file_id, name


def download_drive_file(file_id: str, destination: Path) -> Path:
    service = _connect_drive()

    request = service.files().get_media(
        fileId=file_id,
        supportsAllDrives=True,
    )

    destination.parent.mkdir(parents=True, exist_ok=True)

    with destination.open("wb") as f:
        downloader = MediaIoBaseDownload(f, request)
        done = False

        while not done:
            _status, done = downloader.next_chunk()

    return destination


def upload_drive_file(file_id: str, path: Path) -> None:
    service = _connect_drive()

    media = MediaFileUpload(
        str(path),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        resumable=False,
    )

    service.files().update(
        fileId=file_id,
        media_body=media,
        supportsAllDrives=True,
    ).execute()


def _find_header_row_and_map(ws) -> tuple[int, dict[str, int]]:
    for row_idx in range(1, min(ws.max_row, 25) + 1):
        values = [ws.cell(row=row_idx, column=col).value for col in range(1, ws.max_column + 1)]
        normalized = [_norm(v) for v in values]

        if any(v in normalized for v in {"address", "property", "propertyaddress", "streetaddress"}):
            headers: dict[str, int] = {}

            for col_idx, raw_header in enumerate(values, start=1):
                key = _norm(raw_header)

                if key:
                    headers[key] = col_idx

            return row_idx, headers

    raise ValueError("Could not find header row in BLU Active Deals workbook")


def _find_col(headers: dict[str, int], candidates: list[str]) -> int | None:
    normalized = [_norm(c) for c in candidates]

    for candidate in normalized:
        if candidate in headers:
            return headers[candidate]

    for key, col_idx in headers.items():
        if any(candidate in key for candidate in normalized):
            return col_idx

    return None

def _prepare_active_deals_text(body_text: str) -> str:
    text = (body_text or "").replace("\r\n", "\n").replace("\r", "\n").strip()

    # Some Gmail/iPhone HTML bodies arrive with the address and next label glued:
    #   2607 poplar, wichitaBlu rent est: $925
    #   257 Poplar, wichitaOffer status: BLU2 purchased
    #
    # Insert a newline before known labels even when glued to the prior text.
    known_labels = [
        "address",
        "property address",
        "street address",
        "property",
        "status",
        "status update",
        "offer status",
        "rent",
        "monthly rent",
        "zest rent",
        "market rent",
        "blu rent est",
        "blue rent est",
        "price",
        "asking price",
        "purchase price",
        "offer price",
        "beds",
        "bedrooms",
        "baths",
        "bathrooms",
        "source",
        "notes",
        "note",
        "comments",
        "comment",
    ]

    label_pattern = "|".join(re.escape(label) for label in known_labels)

    text = re.sub(
        rf"(?i)(?<!^)(?<!\n)({label_pattern})\s*:",
        r"\n\1:",
        text,
    )

    return text

def _extract_labeled_fields(body_text: str) -> dict[str, str]:
    fields: dict[str, str] = {}

    text = _prepare_active_deals_text(body_text)

    pattern = re.compile(r"(?im)^\s*([A-Za-z][A-Za-z0-9 /&()._-]{1,50})\s*:\s*(.*)$")
    matches = list(pattern.finditer(text))

    for i, match in enumerate(matches):
        label = match.group(1).strip()
        start_value = match.group(2).strip()

        next_start = matches[i + 1].start() if i + 1 < len(matches) else len(text)
        value_tail_start = match.end()
        tail = text[value_tail_start:next_start].strip()

        if tail:
            value = f"{start_value}\n{tail}".strip()
        else:
            value = start_value

        value = re.sub(r"\n+", " ", value)
        value = re.sub(r"\s+", " ", value).strip()

        if label:
            fields[_norm(label)] = value

    return fields


def _looks_like_address_line(line: str) -> bool:
    line = str(line or "").strip()

    if not line:
        return False

    if ":" in line:
        return False

    return bool(
        re.search(
            r"^\s*\d{2,6}\s+(?:[NSEW]\.?\s+|North\s+|South\s+|East\s+|West\s+)?[A-Za-z0-9.'-]+(?:\s+[A-Za-z0-9.'-]+){0,4}(?:,\s*[A-Za-z .'-]+)?",
            line,
            re.IGNORECASE,
        )
    )


def _extract_address_from_body(body_text: str, fields: dict[str, str]) -> str:
    for key in ["address", "propertyaddress", "streetaddress", "property"]:
        value = fields.get(key)

        if value:
            return value

    text = (body_text or "").replace("\r\n", "\n").replace("\r", "\n")

    for line in text.splitlines():
        line = line.strip()

        if _looks_like_address_line(line):
            return line

    return ""

def parse_active_deal_blocks(body_text: str) -> list[str]:
    """
    Split one Active Deals email into one or more deal blocks.

    Supports:
      2607 poplar, wichita
      Blu rent est: $925

      257 Poplar, wichita
      Offer status: BLU2 purchased

    Also supports:
      Address: 2607 N Poplar St, Wichita
      Rent: 925
      Status: Research
    """

    text = _prepare_active_deals_text(body_text)

    if not text:
        return []

    lines = text.splitlines()

    blocks: list[list[str]] = []
    current: list[str] = []

    for raw_line in lines:
        line = raw_line.strip()

        if not line:
            if current:
                current.append("")
            continue

        is_address = _looks_like_address_line(line) or line.lower().startswith("address:")

        if is_address and current:
            cleaned_current = "\n".join(current).strip()

            if cleaned_current:
                blocks.append(current)

            current = [line]
        else:
            current.append(line)

    if current:
        cleaned_current = "\n".join(current).strip()

        if cleaned_current:
            blocks.append(current)

    return ["\n".join(block).strip() for block in blocks if "\n".join(block).strip()]


def _field_aliases_for_header(header_key: str) -> list[str]:
    aliases = {
        "address": ["address", "property", "propertyaddress", "streetaddress"],
        "property": ["property", "address", "propertyaddress", "streetaddress"],
        "propertyaddress": ["propertyaddress", "address", "property"],
        "city": ["city"],
        "state": ["state"],
        "zip": ["zip", "zipcode", "postalcode"],
        "zipcode": ["zipcode", "zip", "postalcode"],
        "status": ["status", "statusupdate", "offerstatus"],
        "notes": ["notes", "note", "comments", "comment"],
        "rent": ["rent", "monthlyrent", "zestrent", "marketrent", "blurentest", "bluerentest", "bluentest"],
        "monthlyrent": ["monthlyrent", "rent", "zestrent", "marketrent", "blurentest", "bluerentest", "bluentest"],
        "zestrent": ["zestrent", "rent", "monthlyrent"],
        "price": ["price", "askingprice", "purchaseprice", "offerprice"],
        "askingprice": ["askingprice", "price", "purchaseprice"],
        "beds": ["beds", "bedrooms", "br"],
        "bedrooms": ["bedrooms", "beds", "br"],
        "baths": ["baths", "bathrooms", "ba"],
        "bathrooms": ["bathrooms", "baths", "ba"],
        "source": ["source", "leadsource"],
    }

    return aliases.get(header_key, [header_key])


def _split_city_state_zip(address: str) -> tuple[str, str, str]:
    parts = [p.strip() for p in str(address or "").split(",")]

    city = ""
    state = ""
    zip_code = ""

    if len(parts) >= 2:
        city = parts[1].strip()

    if len(parts) >= 3:
        state_zip = parts[2].strip()
        tokens = state_zip.split()

        if tokens:
            state = tokens[0].strip()

        if len(tokens) > 1:
            zip_code = tokens[1].strip()

    return city, state, zip_code


def _existing_match_row(ws, header_row: int, address_col: int, incoming_key: str) -> int | None:
    for row_idx in range(header_row + 1, ws.max_row + 1):
        existing_address = ws.cell(row=row_idx, column=address_col).value
        existing_key = normalize_address_match_key(str(existing_address or ""))

        if existing_key and existing_key == incoming_key:
            return row_idx

    return None


def _update_one_active_deal_block(
    body_text: str,
    file_id: str,
    file_name: str,
    local_path: Path,
) -> ActiveDealsResult:
    fields = _extract_labeled_fields(body_text)
    address = _extract_address_from_body(body_text, fields)

    if not address:
        raise ValueError("Active Deals email did not contain a usable address")

    match_key = normalize_address_match_key(address)

    if not match_key:
        raise ValueError(f"Could not build address match key for: {address}")

    wb = load_workbook(local_path)
    ws = wb.active

    header_row, headers = _find_header_row_and_map(ws)

    address_col = _find_col(headers, ["Address", "Property Address", "Property", "Street Address"])

    if address_col is None:
        raise ValueError("BLU Active Deals workbook does not have an Address/Property column")

    row_idx = _existing_match_row(ws, header_row, address_col, match_key)
    action = "updated"

    if row_idx is None:
        row_idx = ws.max_row + 1
        action = "added"

    city, state, zip_code = _split_city_state_zip(address)

    updated_fields: list[str] = []

    ws.cell(row=row_idx, column=address_col, value=address)
    updated_fields.append("Address")

    city_col = _find_col(headers, ["City"])
    state_col = _find_col(headers, ["State"])
    zip_col = _find_col(headers, ["Zip", "Zip Code", "Postal Code"])

    if city and city_col:
        ws.cell(row=row_idx, column=city_col, value=city)
        updated_fields.append("City")

    if state and state_col:
        ws.cell(row=row_idx, column=state_col, value=state)
        updated_fields.append("State")

    if zip_code and zip_col:
        ws.cell(row=row_idx, column=zip_col, value=zip_code)
        updated_fields.append("Zip")

    for header_key, col_idx in headers.items():
        if col_idx == address_col:
            continue

        aliases = _field_aliases_for_header(header_key)
        value = None

        for alias in aliases:
            if _norm(alias) in fields:
                value = fields[_norm(alias)]
                break

        if value is None or value == "":
            continue

        ws.cell(row=row_idx, column=col_idx, value=value)
        updated_fields.append(str(ws.cell(row=header_row, column=col_idx).value or header_key))

    timestamp_col = _find_col(
        headers,
        ["Last Updated", "Updated", "Last Updated UTC", "Updated UTC"],
    )

    if timestamp_col:
        ws.cell(row=row_idx, column=timestamp_col, value=datetime.utcnow().isoformat())
        updated_fields.append("Last Updated")

    wb.save(local_path)

    return ActiveDealsResult(
        action=action,
        file_name=file_name,
        row_number=row_idx,
        address=address,
        match_key=match_key,
        updated_fields=sorted(set(updated_fields)),
    )


def update_active_deals_from_email(body_text: str) -> list[ActiveDealsResult]:
    blocks = parse_active_deal_blocks(body_text)

    if not blocks:
        blocks = [body_text]

    file_id, file_name = find_latest_active_deals_file()
    results: list[ActiveDealsResult] = []

    with tempfile.TemporaryDirectory(prefix="active_deals_") as tmp:
        local_path = Path(tmp) / file_name
        download_drive_file(file_id, local_path)

        for block in blocks:
            result = _update_one_active_deal_block(
                body_text=block,
                file_id=file_id,
                file_name=file_name,
                local_path=local_path,
            )
            results.append(result)

        upload_drive_file(file_id, local_path)

    return results